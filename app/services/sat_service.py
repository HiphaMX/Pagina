import base64
import datetime
import io
import os
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import List, Optional
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from dotenv import load_dotenv

load_dotenv()

from cfdiclient import Fiel, Autenticacion, DescargaMasiva, VerificaSolicitudDescarga
from cfdiclient import solicitadescargaEmitidos, solicitadescargaRecibidos
from app.models.sat import SatAccount, SatInvoice, SatDownloadRequest

# Definir directorio de secretos relativo a la raíz del proyecto
SECRETS_DIR = Path(__file__).resolve().parent.parent.parent / ".secrets"


def get_fiel_for_rfc(rfc: str) -> Fiel:
    """
    Carga de forma segura el certificado (.cer) y llave (.key) para un RFC
    desde la carpeta local .secrets/ y su contraseña desde el entorno.
    """
    rfc_dir = SECRETS_DIR / rfc.upper()
    if not rfc_dir.exists() or not rfc_dir.is_dir():
        raise Exception(f"No se encontró la carpeta de credenciales para el RFC {rfc} en .secrets/")

    cer_files = list(rfc_dir.glob("*.cer"))
    key_files = list(rfc_dir.glob("*.key"))

    if not cer_files or not key_files:
        raise Exception(f"Falta archivo .cer o .key en el directorio {rfc_dir}")

    cer_path = cer_files[0]
    key_path = key_files[0]

    env_var_name = f"SAT_PASSWORD_{rfc.upper()}"
    password = os.getenv(env_var_name)
    if not password:
        raise Exception(f"No se encontró la contraseña de e.firma en la variable de entorno {env_var_name}")

    with open(cer_path, "rb") as f:
        cer_der = f.read()
    with open(key_path, "rb") as f:
        key_der = f.read()

    return Fiel(cer_der, key_der, password)


def parse_cfdi_xml(xml_content: str) -> Optional[dict]:
    """
    Parsea de manera robusta y en puro Python un XML de CFDI (v3.3 o v4.0).
    Extrae UUID, emisor, receptor, totales e impuestos detallados.
    """
    try:
        root = ET.fromstring(xml_content.encode("utf-8"))
    except Exception as e:
        print(f"Error parsing XML: {e}")
        return None

    # Helper para obtener atributos de forma flexible
    def get_attr(elem, attr_name):
        if elem is None:
            return None
        return elem.attrib.get(attr_name) or elem.attrib.get(attr_name.lower()) or elem.attrib.get(attr_name.capitalize())

    fecha_str = get_attr(root, 'Fecha')
    subtotal = float(get_attr(root, 'SubTotal') or get_attr(root, 'subtotal') or 0.0)
    total = float(get_attr(root, 'Total') or get_attr(root, 'total') or 0.0)
    tipo_comprobante = get_attr(root, 'TipoDeComprobante')

    # Parsear Fecha a objeto datetime
    fecha_dt = None
    if fecha_str:
        try:
            # Los formatos del SAT suelen ser YYYY-MM-DDTHH:MM:SS
            fecha_dt = datetime.datetime.fromisoformat(fecha_str)
        except Exception:
            try:
                fecha_dt = datetime.datetime.strptime(fecha_str, "%Y-%m-%dT%H:%M:%S")
            except Exception:
                fecha_dt = datetime.datetime.now()

    emisor_elem = None
    receptor_elem = None
    tfd_elem = None

    for elem in root.iter():
        tag_local = elem.tag.split('}')[-1]
        if tag_local == 'Emisor':
            emisor_elem = elem
        elif tag_local == 'Receptor':
            receptor_elem = elem
        elif tag_local == 'TimbreFiscalDigital':
            tfd_elem = elem

    uuid = get_attr(tfd_elem, 'UUID')
    if not uuid:
        # Buscar en todo el XML como fallback
        for elem in root.iter():
            tag_local = elem.tag.split('}')[-1]
            if tag_local == 'TimbreFiscalDigital':
                uuid = get_attr(elem, 'UUID')
                break

    if not uuid:
        return None

    emisor_rfc = get_attr(emisor_elem, 'Rfc')
    emisor_nombre = get_attr(emisor_elem, 'Nombre')
    receptor_rfc = get_attr(receptor_elem, 'Rfc')
    receptor_nombre = get_attr(receptor_elem, 'Nombre')

    # Desglose de impuestos
    iva_trasladado = 0.0
    iva_retenido = 0.0
    isr_retenido = 0.0

    for elem in root.iter():
        tag_local = elem.tag.split('}')[-1]
        if tag_local == 'Traslado':
            impuesto = get_attr(elem, 'Impuesto')
            if impuesto == '002':  # 002 = IVA
                iva_trasladado += float(get_attr(elem, 'Importe') or get_attr(elem, 'importe') or 0.0)
        elif tag_local == 'Retencion':
            impuesto = get_attr(elem, 'Impuesto')
            importe = float(get_attr(elem, 'Importe') or get_attr(elem, 'importe') or 0.0)
            if impuesto == '002':  # IVA Retenido
                iva_retenido += importe
            elif impuesto == '001':  # ISR Retenido
                isr_retenido += importe

    # Resumen de conceptos
    conceptos = []
    for elem in root.iter():
        tag_local = elem.tag.split('}')[-1]
        if tag_local == 'Concepto':
            descripcion = get_attr(elem, 'Descripcion')
            cantidad = get_attr(elem, 'Cantidad')
            if descripcion:
                conceptos.append(f"{cantidad or 1}x {descripcion}")
    conceptos_resumen = ", ".join(conceptos)[:255]

    return {
        'uuid': uuid,
        'emisor_rfc': emisor_rfc,
        'emisor_nombre': emisor_nombre,
        'receptor_rfc': receptor_rfc,
        'receptor_nombre': receptor_nombre,
        'fecha_emision': fecha_dt or datetime.datetime.now(),
        'tipo_comprobante': tipo_comprobante,
        'subtotal': subtotal,
        'total': total,
        'iva_trasladado': iva_trasladado,
        'iva_retenido': iva_retenido,
        'isr_retenido': isr_retenido,
        'conceptos_resumen': conceptos_resumen
    }


class SatService:
    @staticmethod
    def solicitar_descarga(db: Session, rfc: str, fecha_inicio: datetime.date, fecha_fin: datetime.date, tipo: str) -> dict:
        """
        Envía una solicitud de descarga masiva al SAT para un RFC y rango de fechas.
        `tipo` puede ser "emitidas" o "recibidas".
        """
        # Convertir a datetime al inicio y fin del día
        f_inicio_dt = datetime.datetime.combine(fecha_inicio, datetime.time.min)
        f_fin_dt = datetime.datetime.combine(fecha_fin, datetime.time.max)

        fiel = get_fiel_for_rfc(rfc)
        auth = Autenticacion(fiel, timeout=60)
        token = auth.obtener_token()

        if tipo == "emitidas":
            descargador = solicitadescargaEmitidos.SolicitaDescargaEmitidos(fiel, timeout=60)
            # En emitidas, el solicitante es el emisor
            resultado = descargador.solicitar_descarga(
                token=token,
                rfc_solicitante=rfc.upper(),
                fecha_inicial=f_inicio_dt,
                fecha_final=f_fin_dt,
                rfc_emisor=rfc.upper(),
                tipo_solicitud='CFDI'
            )
        elif tipo == "recibidas":
            descargador = solicitadescargaRecibidos.SolicitaDescargaRecibidos(fiel, timeout=60)
            # En recibidas, el solicitante es el receptor
            resultado = descargador.solicitar_descarga(
                token=token,
                rfc_solicitante=rfc.upper(),
                fecha_inicial=f_inicio_dt,
                fecha_final=f_fin_dt,
                rfc_receptor=rfc.upper(),
                tipo_solicitud='CFDI'
            )
        else:
            raise ValueError("El tipo de descarga debe ser 'emitidas' o 'recibidas'")

        id_solicitud = resultado.get("id_solicitud")
        cod_estatus = resultado.get("cod_estatus")
        mensaje = resultado.get("mensaje")

        if not id_solicitud:
            raise Exception(f"El SAT rechazó la solicitud. Código: {cod_estatus}. Mensaje: {mensaje}")

        # Guardar solicitud en BD
        db_request = SatDownloadRequest(
            account_rfc=rfc.upper(),
            id_solicitud=id_solicitud,
            fecha_inicio=f_inicio_dt,
            fecha_fin=f_fin_dt,
            tipo_cfdi=tipo,
            status="pendiente"
        )
        db.add(db_request)
        db.commit()
        db.refresh(db_request)

        return {
            "id_solicitud": id_solicitud,
            "status": "pendiente",
            "cod_estatus": cod_estatus,
            "mensaje": mensaje
        }

    @staticmethod
    def verificar_peticiones_pendientes(db: Session) -> dict:
        """
        Consulta en el SAT el estatus de todas las peticiones con estatus 'pendiente'.
        Si ya están listas ('terminada'), descarga el zip y procesa los XMLs.
        """
        pendientes = db.query(SatDownloadRequest).filter(
            SatDownloadRequest.status.in_(["pendiente", "error"]),
            SatDownloadRequest.intentos < 5
        ).all()

        resumen = {"procesadas": 0, "descargadas": 0, "errores": 0}

        import time
        for request in pendientes:
            try:
                # Retardo de 2 segundos para evitar bloqueos por tasa de peticiones (rate limiting) del SAT
                time.sleep(2.0)
                fiel = get_fiel_for_rfc(request.account_rfc)
                auth = Autenticacion(fiel, timeout=60)
                
                # Autenticar con reintentos para mitigar caídas de conexión del SAT
                token = None
                for attempt in range(3):
                    try:
                        token = auth.obtener_token()
                        break
                    except Exception as token_err:
                        if attempt == 2:
                            raise token_err
                        time.sleep(3.0)

                # Verificar descarga con reintentos para mitigar caídas de conexión del SAT
                verificador = VerificaSolicitudDescarga(fiel, timeout=60)
                verificacion = None
                for attempt in range(3):
                    try:
                        verificacion = verificador.verificar_descarga(
                            token=token,
                            rfc_solicitante=request.account_rfc,
                            id_solicitud=request.id_solicitud
                        )
                        break
                    except Exception as verify_err:
                        if attempt == 2:
                            raise verify_err
                        time.sleep(3.0)

                estado_solicitud = verificacion.get("estado_solicitud")
                
                # Convertir a int si es string
                if estado_solicitud is not None:
                    estado_solicitud = int(estado_solicitud)

                if estado_solicitud == 3:  # 3 = Terminada
                    paquetes = verificacion.get("paquetes", [])
                    descargador = DescargaMasiva(fiel, timeout=60)

                    for id_paquete in paquetes:
                        descarga_res = descargador.descargar_paquete(
                            token=token,
                            rfc_solicitante=request.account_rfc,
                            id_paquete=id_paquete
                        )
                        
                        paquete_b64 = descarga_res.get("paquete_b64")
                        if paquete_b64:
                            zip_content = base64.b64decode(paquete_b64)
                            SatService.procesar_paquete_zip(db, request.account_rfc, request.tipo_cfdi, zip_content)
                            resumen["descargadas"] += 1

                    request.status = "descargada"
                    request.intentos = 0  # Reset on success
                    request.mensaje_error = None
                    db.commit()
                    resumen["procesadas"] += 1

                elif estado_solicitud in [1, 2]:  # Aceptada o En Proceso
                    # Sigue pendiente en el SAT, no incrementamos intentos de error
                    request.intentos = 0  # Reset error counters because connection was successful
                    request.status = "pendiente"
                    request.mensaje_error = None
                    db.commit()
                else:
                    # Algún error o rechazada por el SAT
                    request.status = "error"
                    request.intentos += 1
                    request.mensaje_error = f"Estatus del SAT no soportado: {estado_solicitud}"
                    db.commit()
                    resumen["errores"] += 1

            except Exception as e:
                request.status = "error"
                request.intentos += 1
                request.mensaje_error = str(e)
                db.commit()
                resumen["errores"] += 1

        return resumen

    @staticmethod
    def procesar_paquete_zip(db: Session, account_rfc: str, tipo_cfdi: str, zip_content: bytes):
        """
        Descomprime un archivo ZIP de facturas, extrae el contenido XML
        y lo guarda parseado en la base de datos.
        """
        # El tipo_cfdi de la solicitud es plural ("emitidas"/"recibidas"),
        # pero en la factura guardamos singular ("emitida"/"recibida")
        singular_tipo = "emitida" if tipo_cfdi.startswith("emit") else "recibida"

        with zipfile.ZipFile(io.BytesIO(zip_content)) as z:
            for filename in z.namelist():
                if filename.lower().endswith(".xml"):
                    try:
                        xml_bytes = z.read(filename)
                        xml_text = xml_bytes.decode("utf-8", errors="ignore")
                        
                        parsed = parse_cfdi_xml(xml_text)
                        if parsed:
                            # Verificar si ya existe por UUID
                            db_invoice = db.query(SatInvoice).filter(SatInvoice.uuid == parsed['uuid']).first()
                            if not db_invoice:
                                db_invoice = SatInvoice(
                                    account_rfc=account_rfc.upper(),
                                    uuid=parsed['uuid'],
                                    emisor_rfc=parsed['emisor_rfc'].upper() if parsed['emisor_rfc'] else "",
                                    emisor_nombre=parsed['emisor_nombre'],
                                    receptor_rfc=parsed['receptor_rfc'].upper() if parsed['receptor_rfc'] else "",
                                    receptor_nombre=parsed['receptor_nombre'],
                                    fecha_emision=parsed['fecha_emision'],
                                    tipo_comprobante=parsed['tipo_comprobante'],
                                    subtotal=parsed['subtotal'],
                                    total=parsed['total'],
                                    iva_trasladado=parsed['iva_trasladado'],
                                    iva_retenido=parsed['iva_retenido'],
                                    isr_retenido=parsed['isr_retenido'],
                                    tipo_cfdi=singular_tipo,
                                    conceptos_resumen=parsed['conceptos_resumen'],
                                    xml_content=xml_text
                                )
                                db.add(db_invoice)
                            else:
                                # Si ya existe, actualizamos los campos clave
                                db_invoice.emisor_nombre = parsed['emisor_nombre']
                                db_invoice.receptor_nombre = parsed['receptor_nombre']
                                db_invoice.subtotal = parsed['subtotal']
                                db_invoice.total = parsed['total']
                                db_invoice.iva_trasladado = parsed['iva_trasladado']
                                db_invoice.iva_retenido = parsed['iva_retenido']
                                db_invoice.isr_retenido = parsed['isr_retenido']
                                db_invoice.conceptos_resumen = parsed['conceptos_resumen']
                                db_invoice.xml_content = xml_text
                            
                            db.commit()
                    except Exception as e:
                        print(f"Error procesando XML {filename}: {e}")

    @staticmethod
    def generar_reporte_excel(invoices: List[SatInvoice], filepath: str, rfc: str, month_str: str):
        """
        Genera un archivo Excel (.xlsx) estructurado y con un diseño estético premium (Google HSL colors)
        que resume las facturas emitidas y recibidas.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Fiscal"

        # Colores y fuentes (Azul oscuro premium #0D2C54 y grises suaves)
        font_title = Font(name='Arial', size=15, bold=True, color='0D2C54')
        font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        font_data = Font(name='Arial', size=9)
        font_total = Font(name='Arial', size=9, bold=True)

        fill_header = PatternFill(start_color='0D2C54', end_color='0D2C54', fill_type='solid')
        fill_total = PatternFill(start_color='EAF2F8', end_color='EAF2F8', fill_type='solid')

        border_thin = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )
        border_double = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='double', color='000000')
        )

        # Título
        ws.merge_cells('A1:J1')
        ws['A1'] = f"REPORTE CONTABLE DE FACTURACIÓN SAT ({month_str})"
        ws['A1'].font = font_title
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Subtítulo de información
        ws.merge_cells('A2:J2')
        ws['A2'] = f"RFC Contribuyente: {rfc.upper()} | Fecha de generación: {datetime.date.today().strftime('%Y-%m-%d')}"
        ws['A2'].font = Font(name='Arial', size=10, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Encabezados de tabla
        headers = [
            "Tipo CFDI", "Folio Fiscal (UUID)", "RFC Emisor", "Nombre Emisor",
            "RFC Receptor", "Nombre Receptor", "Fecha Emisión", "Subtotal",
            "IVA Trasladado", "Total"
        ]

        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[4].height = 25

        row_idx = 5
        sum_subtotal = 0.0
        sum_iva = 0.0
        sum_total = 0.0

        for inv in invoices:
            ws.cell(row=row_idx, column=1, value=inv.tipo_cfdi.capitalize()).font = font_data
            ws.cell(row=row_idx, column=2, value=inv.uuid).font = font_data
            ws.cell(row=row_idx, column=3, value=inv.emisor_rfc).font = font_data
            ws.cell(row=row_idx, column=4, value=inv.emisor_nombre or "").font = font_data
            ws.cell(row=row_idx, column=5, value=inv.receptor_rfc).font = font_data
            ws.cell(row=row_idx, column=6, value=inv.receptor_nombre or "").font = font_data

            fecha_val = inv.fecha_emision.strftime("%Y-%m-%d %H:%M") if inv.fecha_emision else ""
            ws.cell(row=row_idx, column=7, value=fecha_val).font = font_data

            sub = inv.subtotal or 0.0
            iva = inv.iva_trasladado or 0.0
            tot = inv.total or 0.0

            sum_subtotal += sub
            sum_iva += iva
            sum_total += tot

            c_sub = ws.cell(row=row_idx, column=8, value=sub)
            c_sub.font = font_data
            c_sub.number_format = '$#,##0.00'

            c_iva = ws.cell(row=row_idx, column=9, value=iva)
            c_iva.font = font_data
            c_iva.number_format = '$#,##0.00'

            c_tot = ws.cell(row=row_idx, column=10, value=tot)
            c_tot.font = font_data
            c_tot.number_format = '$#,##0.00'

            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).border = border_thin

            row_idx += 1

        # Fila de Totales
        ws.cell(row=row_idx, column=1, value="TOTALES").font = font_total
        ws.cell(row=row_idx, column=1).fill = fill_total

        for col in range(2, 8):
            ws.cell(row=row_idx, column=col).fill = fill_total

        c_sub_tot = ws.cell(row=row_idx, column=8, value=sum_subtotal)
        c_sub_tot.font = font_total
        c_sub_tot.number_format = '$#,##0.00'
        c_sub_tot.fill = fill_total
        c_sub_tot.border = border_double

        c_iva_tot = ws.cell(row=row_idx, column=9, value=sum_iva)
        c_iva_tot.font = font_total
        c_iva_tot.number_format = '$#,##0.00'
        c_iva_tot.fill = fill_total
        c_iva_tot.border = border_double

        c_tot_tot = ws.cell(row=row_idx, column=10, value=sum_total)
        c_tot_tot.font = font_total
        c_tot_tot.number_format = '$#,##0.00'
        c_tot_tot.fill = fill_total
        c_tot_tot.border = border_double

        ws.row_dimensions[row_idx].height = 22

        # Autoajustar anchos de columnas
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format and ('$' in cell.number_format) and isinstance(cell.value, (int, float)):
                    val = f"${cell.value:,.2f}"
                max_len = max(max_len, len(val))
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

        wb.save(filepath)
